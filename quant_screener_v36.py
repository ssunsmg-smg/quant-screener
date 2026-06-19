"""
==============================================================
  퀀트 주식 스크리너 v6.2 (한국 전용 완전통합판)
  ★ v35 전체 기능 유지 + 4대 신규 엔진 추가
  ★ 한국 KOSPI/KOSDAQ 전용 — 미국(S&P500) 별도 운영

  ┌─────────────────────────────────────────────────────────┐
  │  NEW ① 백테스트 엔진   (Walk-forward, IS/OOS 검증)      │
  │  NEW ② 팩터 모델 고도화 (Z-score, 섹터중립, 앙상블)    │
  │  NEW ③ KIS API 자동매매 (한국투자증권 REST, 64bit OK)   │
  │  NEW ④ 실시간 모니터링  (텔레그램 알림, 성과 추적)      │
  └─────────────────────────────────────────────────────────┘

  실행 방법:
    python quant_screener_v36.py              # 대화형 메뉴
    python quant_screener_v36.py --auto       # 자동(스케줄러)
    python quant_screener_v36.py --backtest   # 백테스트 전용
    python quant_screener_v36.py --trade      # KIS 자동매매
    python quant_screener_v36.py --monitor    # 모니터링 전용

  KIS API 설정:
    - https://apiportal.koreainvestment.com 에서 앱키 발급 (무료)
    - kis_config.json 에 app_key / app_secret / account_no 입력
    - 64bit Python, VS Code 환경 그대로 사용 가능

  필요 패키지:
    pip install finance-datareader yfinance requests beautifulsoup4
                pandas numpy openpyxl tqdm dart-fss scikit-learn
==============================================================
"""

import os, sys, re, time, json, zipfile, io, warnings, argparse, traceback, hashlib
import threading
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional

import requests
import pandas as pd
import numpy as np

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    import yfinance as yf
    HAS_YFINANCE = True
except ImportError:
    HAS_YFINANCE = False

try:
    import FinanceDataReader as fdr
    HAS_FDR = True
except ImportError:
    HAS_FDR = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.merge import MergedCell
    from openpyxl.formatting.rule import ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    from sklearn.preprocessing import StandardScaler
    from sklearn.linear_model import Ridge
    HAS_SKLEARN = True
except ImportError:
    HAS_SKLEARN = False

warnings.filterwarnings("ignore")

VERSION   = "v6.1.0"
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(BASE_DIR, ".cache")
LOG_DIR   = os.path.join(BASE_DIR, "logs")
BT_DIR    = os.path.join(BASE_DIR, "backtest")
TRADE_DIR = os.path.join(BASE_DIR, "trades")

for _d in [CACHE_DIR, LOG_DIR, BT_DIR, TRADE_DIR]:
    os.makedirs(_d, exist_ok=True)

# ══════════════════════════════════════════════════════════
# 공통 스타일 상수 (v35 동일)
# ══════════════════════════════════════════════════════════
_HDR_FILL = PatternFill("solid", fgColor="1F3864") if HAS_OPENPYXL else None
_GLD_FILL = PatternFill("solid", fgColor="FFD700") if HAS_OPENPYXL else None
_SLV_FILL = PatternFill("solid", fgColor="C0C0C0") if HAS_OPENPYXL else None
_BRZ_FILL = PatternFill("solid", fgColor="CD7F32") if HAS_OPENPYXL else None
_GRN_FILL = PatternFill("solid", fgColor="E2EFDA") if HAS_OPENPYXL else None
_ALT_FILL = PatternFill("solid", fgColor="F5F5F5") if HAS_OPENPYXL else None
_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
) if HAS_OPENPYXL else None

NAVER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": "https://finance.naver.com",
}


# ══════════════════════════════════════════════════════════
# 공통 유틸 (v35 동일)
# ══════════════════════════════════════════════════════════
def _safe(v, default=None):
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return default
        return v
    except Exception:
        return default

def safe_div(a, b, default=np.nan):
    try:
        if b == 0 or pd.isna(b) or pd.isna(a):
            return default
        return a / b
    except Exception:
        return default

def grade_label(score):
    if score >= 85:   return "◆◆◆ A+ 탁월", "FF4444"
    elif score >= 75: return "◆◆  A  우수",  "FF9900"
    elif score >= 65: return "◆   B  양호",  "FFCC00"
    elif score >= 50: return "    C  보통",   "92D050"
    else:             return "    D  검토",   "BFBFBF"

def _rank_fill(rank):
    if not HAS_OPENPYXL: return None
    if rank == 1:   return _GLD_FILL
    elif rank == 2: return _SLV_FILL
    elif rank == 3: return _BRZ_FILL
    elif rank <= 5: return _GRN_FILL
    else:           return _ALT_FILL

def auto_col_width(ws, min_w=8, max_w=45):
    if not HAS_OPENPYXL: return
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            text = str(cell.value)
            w = sum(2 if ord(c) > 127 else 1 for c in text) + 2
            col_letter = get_column_letter(cell.column)
            col_widths[col_letter] = max(col_widths.get(col_letter, 0), w)
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = min(max(w, min_w), max_w)


# ══════════════════════════════════════════════════════════
# 캐시 시스템 (v35 동일)
# ══════════════════════════════════════════════════════════
_TIER_TTL = {"A": 7, "B": 1}

def _cache_path(key: str, tier: str = "B") -> str:
    safe_key = hashlib.md5(key.encode()).hexdigest()[:12]
    tier_dir = os.path.join(CACHE_DIR, f"tier_{tier}")
    os.makedirs(tier_dir, exist_ok=True)
    return os.path.join(tier_dir, f"{safe_key}.json")

def cache_get(key: str, tier: str = "B"):
    ttl_days = _TIER_TTL.get(tier, 1)
    path = _cache_path(key, tier)
    if not os.path.exists(path):
        return None
    try:
        mtime = os.path.getmtime(path)
        if (time.time() - mtime) / 86400 > ttl_days:
            return None
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def cache_set(key: str, data, tier: str = "B") -> None:
    path = _cache_path(key, tier)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass

def cache_clear(days_old: int = 8) -> None:
    deleted = 0
    cutoff = time.time()
    try:
        for tier_name in ["tier_A", "tier_B", ""]:
            target_dir = os.path.join(CACHE_DIR, tier_name) if tier_name else CACHE_DIR
            if not os.path.isdir(target_dir):
                continue
            for fname in os.listdir(target_dir):
                if not fname.endswith(".json"):
                    continue
                fpath = os.path.join(target_dir, fname)
                try:
                    if (cutoff - os.path.getmtime(fpath)) / 86400 >= days_old:
                        os.remove(fpath)
                        deleted += 1
                except Exception:
                    pass
        if deleted:
            print(f"  [캐시] 오래된 캐시 {deleted}개 삭제")
    except Exception:
        pass

def cache_show_status() -> None:
    for tier in ["A", "B"]:
        tier_dir = os.path.join(CACHE_DIR, f"tier_{tier}")
        if not os.path.isdir(tier_dir):
            continue
        files = [f for f in os.listdir(tier_dir) if f.endswith(".json")]
        ttl = _TIER_TTL[tier]
        valid = sum(
            1 for f in files
            if (time.time() - os.path.getmtime(os.path.join(tier_dir, f))) / 86400 <= ttl
        )
        label = "DART 재무(7일)" if tier == "A" else "네이버 보조(1일)"
        print(f"  💾 캐시 TIER-{tier} [{label}]: {valid}/{len(files)}개 유효")


# ══════════════════════════════════════════════════════════
# NEW ① 백테스트 엔진 (Walk-forward + IS/OOS 검증)
# ══════════════════════════════════════════════════════════
class BacktestEngine:
    """
    Walk-forward 백테스트 엔진 (KOSPI/KOSDAQ 전용)
    - IS(학습) → OOS(검증) 슬라이딩 반복
    - 벤치마크: KOSPI 지수 (^KS11)
    - 과적합 판단: IS/OOS Sharpe 비율 비교
    """

    def __init__(self, universe_codes: list, start: str = "2020-01-01",
                 end: str = None, is_months: int = 12, oos_months: int = 3,
                 top_n: int = 20, initial_capital: float = 100_000_000):
        self.universe   = universe_codes
        self.start      = pd.Timestamp(start)
        self.end        = pd.Timestamp(end) if end else pd.Timestamp.today()
        self.is_months  = is_months
        self.oos_months = oos_months
        self.top_n      = top_n
        self.capital    = initial_capital
        self.results    = {}

    def load_price_data(self) -> pd.DataFrame:
        """yfinance에서 KOSPI/KOSDAQ 종가 로드 (.KS / .KQ 자동 처리)"""
        print(f"  [백테스트] 가격 로드 중... ({len(self.universe)}개 종목)")
        cache_key = f"bt_kr_{hashlib.md5(','.join(self.universe[:15]).encode()).hexdigest()[:8]}"
        cached = cache_get(cache_key, tier="A")
        if cached:
            df = pd.DataFrame(cached)
            df.index = pd.to_datetime(df.index)
            print(f"  [백테스트] 캐시 로드: {df.shape}")
            return df

        if not HAS_YFINANCE:
            print("  ⚠ yfinance 미설치 → 백테스트 불가")
            return pd.DataFrame()

        # KOSDAQ 종목은 .KQ, KOSPI는 .KS (FDR 있으면 시장 구분, 없으면 .KS 시도)
        yf_tickers = [f"{c}.KS" for c in self.universe]
        try:
            raw = yf.download(
                yf_tickers,
                start=self.start.strftime("%Y-%m-%d"),
                end=self.end.strftime("%Y-%m-%d"),
                progress=False, group_by="ticker",
                auto_adjust=True, threads=True
            )
            close_df = pd.DataFrame()
            for code, yf_t in zip(self.universe, yf_tickers):
                try:
                    c_s = (raw["Close"][yf_t] if isinstance(raw.columns, pd.MultiIndex)
                           else raw["Close"]).dropna()
                    if len(c_s) > 50:
                        close_df[code] = c_s
                except Exception:
                    pass
            cache_set(cache_key, close_df.to_dict(), tier="A")
            print(f"  [백테스트] 가격 로드 완료: {close_df.shape}")
            return close_df
        except Exception as e:
            print(f"  ⚠ 가격 로드 실패: {e}")
            return pd.DataFrame()

    def compute_factor_scores(self, price_df: pd.DataFrame,
                               as_of: pd.Timestamp) -> pd.Series:
        """포인트-인-타임 팩터 스코어 (모멘텀 50% + 저변동성 30% + 52주위치 20%)"""
        scores = {}
        for code in price_df.columns:
            try:
                hist = price_df[code][:as_of].dropna()
                if len(hist) < 60:
                    continue
                p_now = hist.iloc[-1]
                p_1m  = hist.iloc[-21] if len(hist) >= 21 else hist.iloc[0]
                p_12m = hist.iloc[-252] if len(hist) >= 252 else hist.iloc[0]
                mom   = (p_1m / p_12m - 1) * 100 if p_12m > 0 else 0
                vol   = hist.pct_change().dropna().tail(60).std() * np.sqrt(252) * 100
                hi52  = hist.tail(252).max()
                lo52  = hist.tail(252).min()
                pos52 = (p_now - lo52) / (hi52 - lo52) * 100 if hi52 > lo52 else 50
                scores[code] = mom * 0.5 + (100 - min(vol, 100)) * 0.3 + pos52 * 0.2
            except Exception:
                pass
        return pd.Series(scores)

    def run_walkforward(self, price_df: pd.DataFrame) -> dict:
        """Walk-forward 검증 실행"""
        print(f"\n  [백테스트] Walk-forward 시작")
        print(f"     기간: {self.start.date()} ~ {self.end.date()}")
        print(f"     IS: {self.is_months}개월 / OOS: {self.oos_months}개월 / 상위: {self.top_n}종목")

        rebal_dates = pd.date_range(
            start=self.start + pd.DateOffset(months=self.is_months),
            end=self.end, freq=f"{self.oos_months}MS"
        )

        portfolio_value = self.capital
        equity_curve    = [{"date": str(self.start.date()), "value": portfolio_value}]
        oos_returns, is_sharpes, period_log = [], [], []

        for rebal_date in rebal_dates:
            is_end   = rebal_date - pd.DateOffset(days=1)
            is_start = is_end - pd.DateOffset(months=self.is_months)
            oos_end  = min(rebal_date + pd.DateOffset(months=self.oos_months) - pd.DateOffset(days=1),
                           self.end)

            scores   = self.compute_factor_scores(price_df, is_end)
            if scores.empty:
                continue
            selected = scores.nlargest(self.top_n).index.tolist()

            oos_rets = self._period_returns(price_df, selected, rebal_date, oos_end)
            if oos_rets is None or len(oos_rets) < 2:
                continue

            port_ret  = oos_rets.mean(axis=1)
            period_r  = float((1 + port_ret).cumprod().iloc[-1] - 1) * 100
            bench_r   = self._benchmark_return(rebal_date, oos_end)
            oos_sharpe= self._sharpe(port_ret)

            is_rets   = self._period_returns(price_df, selected, is_start, is_end)
            is_sharpe = self._sharpe(is_rets.mean(axis=1)) if is_rets is not None else 0

            portfolio_value *= (1 + period_r / 100)
            equity_curve.append({"date": str(oos_end.date()), "value": round(portfolio_value)})
            oos_returns.append(period_r)
            is_sharpes.append(is_sharpe)
            period_log.append({
                "기간":       f"{rebal_date.date()}~{oos_end.date()}",
                "선정종목":   ", ".join(selected[:5]),
                "포트수익(%)":round(period_r, 2),
                "벤치수익(%)":round(bench_r, 2),
                "Alpha(%)":   round(period_r - bench_r, 2),
                "OOS_Sharpe": round(oos_sharpe, 2),
                "IS_Sharpe":  round(is_sharpe, 2),
            })
            print(f"  {rebal_date.date()}~{oos_end.date()}: "
                  f"포트 {period_r:+.1f}% / 벤치 {bench_r:+.1f}% / "
                  f"Alpha {period_r-bench_r:+.1f}% / Sharpe {oos_sharpe:.2f}")

        if not oos_returns:
            print("  ⚠ 유효한 OOS 구간 없음")
            return {}

        years = (self.end - self.start).days / 365.25
        cagr  = ((portfolio_value / self.capital) ** (1 / max(years, 0.1)) - 1) * 100
        all_r = np.array(oos_returns) / 100
        ann_vol = np.std(all_r) * np.sqrt(12 / self.oos_months) * 100
        sharpe  = (np.mean(all_r) * 12/self.oos_months) / max(
                   np.std(all_r) * np.sqrt(12/self.oos_months), 0.001)

        equity_vals = [e["value"] for e in equity_curve]
        peak, mdd = equity_vals[0], 0
        for v in equity_vals:
            peak = max(peak, v)
            mdd  = min(mdd, (v - peak) / peak * 100)

        avg_is  = np.mean(is_sharpes)
        overfit = "⚠ 과적합 의심" if sharpe / max(abs(avg_is), 0.01) < 0.5 else "✅ 정상"

        result = {
            "summary": {
                "기간":        f"{self.start.date()} ~ {self.end.date()}",
                "총수익률(%)": round((portfolio_value / self.capital - 1) * 100, 2),
                "CAGR(%)":     round(cagr, 2),
                "연변동성(%)": round(ann_vol, 2),
                "Sharpe_OOS":  round(sharpe, 2),
                "IS_Sharpe":   round(avg_is, 2),
                "MDD(%)":      round(mdd, 2),
                "승률(%)":     round(sum(1 for r in oos_returns if r > 0) / len(oos_returns) * 100, 1),
                "구간수":      len(oos_returns),
                "과적합판단":  overfit,
                "초기자본(원)":self.capital,
                "최종자본(원)":round(portfolio_value),
            },
            "equity_curve": equity_curve,
            "period_log":   period_log,
        }
        self.results = result
        self._print_summary(result)
        return result

    def _period_returns(self, price_df, codes, start, end):
        try:
            avail = [c for c in codes if c in price_df.columns]
            if not avail:
                return None
            sub = price_df[avail].loc[start:end].dropna(how="all")
            if len(sub) < 2:
                return None
            return sub.pct_change().dropna(how="all")
        except Exception:
            return None

    def _benchmark_return(self, start, end) -> float:
        try:
            if not HAS_YFINANCE:
                return 0.0
            h = yf.download("^KS11",
                            start=start.strftime("%Y-%m-%d"),
                            end=(end + timedelta(days=5)).strftime("%Y-%m-%d"),
                            progress=False, auto_adjust=True)
            if h.empty or len(h) < 2:
                return 0.0
            c = h["Close"].dropna()
            return float(c.iloc[-1] / c.iloc[0] - 1) * 100
        except Exception:
            return 0.0

    def _sharpe(self, returns_series, rf=0.03) -> float:
        try:
            if returns_series is None or len(returns_series) < 2:
                return 0.0
            r  = np.array(returns_series, dtype=float)
            ex = r - rf / 252
            return float(np.mean(ex) / np.std(ex) * np.sqrt(252)) if np.std(ex) > 1e-10 else 0.0
        except Exception:
            return 0.0

    def _print_summary(self, result: dict):
        s = result.get("summary", {})
        print("\n" + "=" * 65)
        print(f"  📊 백테스트 결과 [{s.get('기간','')}]")
        print("=" * 65)
        for k, v in s.items():
            if k in ("초기자본(원)", "최종자본(원)"):
                print(f"  {k:<14}: {int(v):>14,}원")
            elif isinstance(v, float):
                color = "\033[34m" if v > 0 else "\033[31m"
                print(f"  {k:<14}: {color}{v:>+.2f}\033[0m")
            else:
                print(f"  {k:<14}: {v}")
        print("=" * 65)

    def save_to_excel(self, result: dict) -> str:
        if not HAS_OPENPYXL or not result:
            return ""
        fname = datetime.today().strftime("backtest_KR_%Y%m%d_%H%M%S.xlsx")
        fpath = os.path.join(BT_DIR, fname)
        wb = Workbook()
        KR = "맑은 고딕"

        # 시트1: 요약
        ws1 = wb.active
        ws1.title = "백테스트요약"
        ws1.merge_cells("A1:C1")
        h = ws1.cell(1, 1, f"Walk-forward 백테스트 결과  {VERSION}")
        h.font = Font(name=KR, bold=True, size=14, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="1F3864")
        h.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[1].height = 28

        for ri, (k, v) in enumerate(result.get("summary", {}).items(), 2):
            c1 = ws1.cell(ri, 1, k)
            c1.font = Font(name=KR, bold=True, size=10)
            c2 = ws1.cell(ri, 2, v)
            c2.font = Font(name=KR, size=10)
            c2.border = _THIN
            if any(x in k for x in ("수익률", "CAGR", "Sharpe")):
                try:
                    fv = float(v)
                    c2.font = Font(name=KR, size=10, bold=True,
                                   color="0070C0" if fv > 0 else "C00000")
                except Exception:
                    pass

        # 시트2: 구간별 성과
        ws2 = wb.create_sheet("구간별성과")
        plog = result.get("period_log", [])
        if plog:
            hdrs = list(plog[0].keys())
            for j, h in enumerate(hdrs, 1):
                c = ws2.cell(1, j, h)
                c.font  = Font(name=KR, bold=True, color="FFFFFF", size=9)
                c.fill  = PatternFill("solid", fgColor="1F3864")
                c.border= _THIN
                c.alignment = Alignment(horizontal="center")
            for ri, p in enumerate(plog, 2):
                for j, k in enumerate(hdrs, 1):
                    v = p.get(k)
                    c = ws2.cell(ri, j, v)
                    c.font   = Font(name=KR, size=9)
                    c.border = _THIN
                    c.alignment = Alignment(horizontal="center")
                    if k == "Alpha(%)":
                        try:
                            fv = float(v)
                            c.font = Font(name=KR, size=9, bold=True,
                                          color="0070C0" if fv > 0 else "C00000")
                        except Exception:
                            pass

        # 시트3: 수익곡선
        ws3 = wb.create_sheet("수익곡선")
        ws3.cell(1, 1, "날짜").font = Font(name=KR, bold=True)
        ws3.cell(1, 2, "포트폴리오(원)").font = Font(name=KR, bold=True)
        for ri, e in enumerate(result.get("equity_curve", []), 2):
            ws3.cell(ri, 1, e["date"])
            c = ws3.cell(ri, 2, e["value"])
            c.number_format = "#,##0"

        wb.save(fpath)
        print(f"  ✅ 백테스트 엑셀: {fpath}")
        return fpath


# ══════════════════════════════════════════════════════════
# NEW ② 팩터 모델 고도화 (Z-score 섹터중립 + 앙상블)
# ══════════════════════════════════════════════════════════
class FactorModel:
    """
    Z-score 표준화 기반 팩터 모델 (한국 주식 전용)
    - 섹터(업종) 내 Z-score → 업종 편향 제거
    - Winsorize ±3σ → 아웃라이어 클리핑
    - 강화복합점수 = 기존점수 70% + Z팩터 30%
    - 선택적 Ridge ML 가중치 학습
    """

    # 팩터 → 컬럼 매핑 (v35 컬럼명 기준)
    FACTOR_COLS = {
        "value":    ["PBR",   "ROE",    "DIV",    "영업이익률"],
        "momentum": ["52주위치", "6개월수익률", "거래량비율VR(%)"],
        "quality":  ["ROA",   "부채비율", "interest_coverage", "altman_z"],
        "growth":   ["매출성장률", "fcf_margin"],
        "cashflow": ["cfo",   "fcf",    "fcf_margin"],
    }
    # 낮을수록 좋은 팩터 (Z-score 반전 대상)
    INVERT_COLS = {"PBR", "부채비율"}

    def __init__(self, sector_neutral: bool = True, winsorize_sigma: float = 3.0):
        self.sector_neutral  = sector_neutral
        self.winsorize_sigma = winsorize_sigma

    def _winsorize(self, s: pd.Series) -> pd.Series:
        mu, sigma = s.mean(), s.std()
        if sigma < 1e-10:
            return s
        return s.clip(mu - self.winsorize_sigma * sigma,
                      mu + self.winsorize_sigma * sigma)

    def _zscore(self, s: pd.Series) -> pd.Series:
        mu, sigma = s.mean(), s.std()
        if sigma < 1e-10:
            return pd.Series(0.0, index=s.index)
        return (s - mu) / sigma

    def _sector_zscore(self, df: pd.DataFrame, col: str) -> pd.Series:
        """업종 내 Z-score. 업종 종목 수 < 3이면 전체 Z-score 사용"""
        result = pd.Series(np.nan, index=df.index)
        has_sector = "업종" in df.columns and self.sector_neutral

        if has_sector:
            for sector, grp in df.groupby("업종"):
                vals = pd.to_numeric(grp[col], errors="coerce").dropna()
                if len(vals) < 3:
                    ref  = pd.to_numeric(df[col], errors="coerce").dropna()
                    z    = self._zscore(self._winsorize(ref))
                    result.loc[grp.index] = z.reindex(grp.index)
                else:
                    z = self._zscore(self._winsorize(vals))
                    result.loc[vals.index] = z.values
        else:
            vals = pd.to_numeric(df[col], errors="coerce").dropna()
            z    = self._zscore(self._winsorize(vals))
            result.loc[vals.index] = z.values

        return result

    def compute_factor_zscores(self, df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        group_scores = {}

        for group, cols in self.FACTOR_COLS.items():
            group_z = []
            for col in cols:
                if col not in df.columns:
                    continue
                z = self._sector_zscore(df, col)
                z_col = f"z_{col}"
                out[z_col] = -z if col in self.INVERT_COLS else z
                group_z.append(out[z_col])

            if group_z:
                out[f"factor_{group}"] = pd.concat(group_z, axis=1).mean(axis=1)
                group_scores[group] = f"factor_{group}"

        avail = [v for v in group_scores.values() if v in out.columns]
        if avail:
            out["factor_composite"]  = out[avail].mean(axis=1)
            out["factor_score_pct"]  = (out["factor_composite"].rank(pct=True) * 100).round(1)

        return out

    def ml_signal(self, df: pd.DataFrame, target_col: str = "6개월수익률") -> pd.Series:
        """Ridge 회귀 팩터 가중치 학습 (백테스트 전용 — 미래 데이터 주의)"""
        if not HAS_SKLEARN:
            print("  ⚠ scikit-learn 미설치 → ML 신호 스킵")
            return pd.Series(dtype=float)
        f_cols = [c for c in df.columns
                  if c.startswith("factor_") and c != "factor_composite"]
        if not f_cols or target_col not in df.columns:
            return pd.Series(dtype=float)
        X  = df[f_cols].fillna(0)
        y  = pd.to_numeric(df[target_col], errors="coerce").fillna(0)
        ok = X.notna().all(axis=1) & y.notna()
        if ok.sum() < 30:
            return pd.Series(dtype=float)
        mdl = Ridge(alpha=1.0).fit(X[ok], y[ok])
        print(f"  [팩터모델] Ridge 학습 ({ok.sum()}개 샘플)")
        for c, w in zip(f_cols, mdl.coef_):
            print(f"     {c.replace('factor_',''):>12}: {w:+.3f}")
        return (pd.Series(mdl.predict(X), index=df.index).rank(pct=True) * 100).round(1)

    def enhance_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """복합점수 강화: 기존 70% + Z팩터 30%"""
        out  = self.compute_factor_zscores(df)
        comp = pd.to_numeric(out.get("복합점수", pd.Series(dtype=float)), errors="coerce").fillna(0)
        zfac = pd.to_numeric(out.get("factor_score_pct", pd.Series(50, index=out.index)),
                             errors="coerce").fillna(50)
        out["강화복합점수"] = (comp * 0.70 + zfac * 0.30).round(1)
        print(f"  [팩터모델] 강화복합점수 계산 완료 (기존 70% + Z팩터 30%)")
        return out


# ══════════════════════════════════════════════════════════
# NEW ③ KIS API 자동매매 엔진
# (한국투자증권 KIS Developers REST API)
#
#  ★ 64bit Python / VS Code 환경 그대로 사용 가능
#  ★ 별도 설치 없음 — requests 라이브러리만 사용
#
#  설정파일: kis_config.json
#  {
#      "app_key":          "발급받은 앱키",
#      "app_secret":       "발급받은 앱시크릿",
#      "account_no":       "12345678-01",  ← 계좌번호-상품코드
#      "is_real":          false,          ← true=실전, false=모의
#      "max_position_pct": 5.0,            ← 종목당 최대 비중(%)
#      "max_daily_trades": 10,             ← 일 최대 거래횟수
#      "stop_loss_pct":    8.0,            ← 개별 손절 기준(%)
#      "take_profit_pct":  20.0            ← 개별 익절 기준(%)
#  }
#
#  앱키 발급: https://apiportal.koreainvestment.com
# ══════════════════════════════════════════════════════════
class KISAutoTrader:
    """
    한국투자증권 KIS Developers API 자동매매 클라이언트
    - REST API (HTTP) 방식 → 64bit Python 완전 지원
    - 실전투자 / 모의투자 구분
    - 매수·매도·잔고·현재가 조회
    - 포지션 관리, 손절·익절 자동 실행
    - 일일 거래한도 안전장치
    """

    BASE_REAL = "https://openapi.koreainvestment.com:9443"
    BASE_MOCK = "https://openapivts.koreainvestment.com:29443"

    def __init__(self, config_path: str = None):
        self.config_path  = config_path or os.path.join(BASE_DIR, "kis_config.json")
        self.cfg          = {}
        self.token        = None
        self.token_exp    = None
        self.is_real      = False
        self.daily_trades = 0
        self.positions    = {}   # {code: {"qty": N, "avg_price": P, "buy_date": D}}
        self._lock        = threading.Lock()
        self._load_config()

    # ── 설정 로드 ──
    def _load_config(self):
        if not os.path.exists(self.config_path):
            self._create_default_config()
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                self.cfg = json.load(f)
            self.is_real = self.cfg.get("is_real", False)
            mode = "🔴 실전투자" if self.is_real else "🟡 모의투자"
            print(f"  [KIS] 설정 로드 완료 ({mode})")
            print(f"  [KIS] 계좌: {self.cfg.get('account_no', '')}")
        except Exception as e:
            print(f"  ⚠ KIS 설정 로드 오류: {e}")

    def _create_default_config(self):
        default = {
            "app_key":          "YOUR_APP_KEY",
            "app_secret":       "YOUR_APP_SECRET",
            "account_no":       "12345678-01",
            "is_real":          False,
            "max_position_pct": 5.0,
            "max_daily_trades": 10,
            "stop_loss_pct":    8.0,
            "take_profit_pct":  20.0,
        }
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        print(f"  [KIS] 기본 설정파일 생성: {self.config_path}")
        print(f"  → https://apiportal.koreainvestment.com 에서 앱키 발급 후 입력")

    @property
    def base_url(self):
        return self.BASE_REAL if self.is_real else self.BASE_MOCK

    def _is_configured(self) -> bool:
        key = self.cfg.get("app_key", "")
        return bool(key and key != "YOUR_APP_KEY")

    # ── OAuth2 토큰 발급 ──
    def get_access_token(self) -> Optional[str]:
        """액세스 토큰 발급 (유효기간 24시간, 자동 캐시)"""
        if self.token and self.token_exp and datetime.now() < self.token_exp:
            return self.token
        if not self._is_configured():
            print("  ⚠ KIS 앱키 미설정 → kis_config.json 확인")
            return None
        try:
            resp = requests.post(
                f"{self.base_url}/oauth2/tokenP",
                json={
                    "grant_type": "client_credentials",
                    "appkey":     self.cfg["app_key"],
                    "appsecret":  self.cfg["app_secret"],
                },
                timeout=10
            )
            data = resp.json()
            if "access_token" in data:
                self.token     = data["access_token"]
                self.token_exp = datetime.now() + timedelta(hours=23)
                print(f"  [KIS] 토큰 발급 완료 (유효: {self.token_exp.strftime('%H:%M')}까지)")
                return self.token
            else:
                print(f"  ⚠ KIS 토큰 발급 실패: {data.get('msg1', '')}")
                return None
        except Exception as e:
            print(f"  ⚠ KIS 토큰 요청 오류: {e}")
            return None

    def _headers(self, tr_id: str) -> dict:
        """공통 API 요청 헤더"""
        token = self.get_access_token()
        return {
            "Content-Type":  "application/json; charset=utf-8",
            "authorization": f"Bearer {token}",
            "appkey":        self.cfg.get("app_key", ""),
            "appsecret":     self.cfg.get("app_secret", ""),
            "tr_id":         tr_id,
            "custtype":      "P",
        }

    # ── 잔고 조회 ──
    def get_balance(self) -> dict:
        """주식 잔고 조회 (총평가금액·예수금·순자산)"""
        if not self._is_configured():
            return {"총평가금액": 0, "예수금총액": 0, "순자산": 10_000_000, "holdings": []}

        tr_id   = "TTTC8434R" if self.is_real else "VTTC8434R"
        acc     = self.cfg.get("account_no", "")
        acc_no  = acc[:8] if len(acc) >= 8 else acc
        acc_prd = acc[9:] if len(acc) > 9 else "01"

        try:
            resp = requests.get(
                f"{self.base_url}/uapi/domestic-stock/v1/trading/inquire-balance",
                headers=self._headers(tr_id),
                params={
                    "CANO": acc_no, "ACNT_PRDT_CD": acc_prd,
                    "AFHR_FLPR_YN": "N", "OFL_YN": "",
                    "INQR_DVSN": "02", "UNPR_DVSN": "01",
                    "FUND_STTL_ICLD_YN": "N", "FNCG_AMT_AUTO_RDPT_YN": "N",
                    "PRCS_DVSN": "01", "CTX_AREA_FK100": "", "CTX_AREA_NK100": "",
                },
                timeout=10
            )
            data = resp.json()
            if data.get("rt_cd") == "0":
                out2    = data.get("output2", [{}])
                bal     = out2[0] if out2 else {}
                result  = {
                    "총평가금액": int(bal.get("tot_evlu_amt", 0)),
                    "예수금총액": int(bal.get("dnca_tot_amt", 0)),
                    "순자산":     int(bal.get("nass_amt", 0)),
                    "holdings":   data.get("output1", []),
                }
                print(f"  [KIS] 잔고: 총평가 {result['총평가금액']:,}원 / "
                      f"예수금 {result['예수금총액']:,}원")
                return result
            else:
                print(f"  ⚠ KIS 잔고 조회 실패: {data.get('msg1', '')}")
                return {"총평가금액": 0, "예수금총액": 0, "순자산": 10_000_000, "holdings": []}
        except Exception as e:
            print(f"  ⚠ KIS 잔고 조회 오류: {e}")
            return {"총평가금액": 0, "예수금총액": 0, "순자산": 10_000_000, "holdings": []}

    # ── 현재가 조회 ──
    def get_current_price(self, stock_code: str) -> int:
        """주식 현재가 조회"""
        try:
            resp = requests.get(
                f"{self.base_url}/uapi/domestic-stock/v1/quotations/inquire-price",
                headers=self._headers("FHKST01010100"),
                params={"FID_COND_MRKT_DIV_CODE": "J", "FID_INPUT_ISCD": stock_code},
                timeout=5
            )
            data = resp.json()
            if data.get("rt_cd") == "0":
                return int(data["output"].get("stck_prpr", 0))
        except Exception:
            pass
        # 폴백: 네이버 금융
        return self._price_naver(stock_code)

    def _price_naver(self, stock_code: str) -> int:
        """네이버 금융 현재가 (KIS API 폴백용)"""
        try:
            url  = f"https://finance.naver.com/item/sise.naver?code={stock_code}"
            resp = requests.get(url, headers=NAVER_HEADERS, timeout=5)
            m    = re.search(r'id="_nowVal"[^>]*>([\d,]+)', resp.text)
            if m:
                return int(m.group(1).replace(",", ""))
        except Exception:
            pass
        return 0

    # ── 주문 실행 ──
    def place_order(self, stock_code: str, order_type: str,
                    qty: int, price: int = 0, reason: str = "") -> dict:
        """
        주문 실행
        order_type : "BUY" | "SELL"
        price      : 0 = 시장가, 양수 = 지정가
        """
        with self._lock:
            max_trades = self.cfg.get("max_daily_trades", 10)
            if self.daily_trades >= max_trades:
                msg = f"  ⚠ 일일 거래한도({max_trades}회) 초과 → 주문 차단: {stock_code}"
                print(msg)
                return {"success": False, "msg": msg}

            if not self._is_configured():
                print("  ⚠ KIS 앱키 미설정 → 주문 불가")
                return {"success": False, "msg": "앱키 미설정"}

            # TR ID (실전 / 모의 구분)
            if self.is_real:
                tr_id = "TTTC0802U" if order_type == "BUY" else "TTTC0801U"
            else:
                tr_id = "VTTC0802U" if order_type == "BUY" else "VTTC0801U"

            acc    = self.cfg.get("account_no", "")
            acc_no = acc[:8] if len(acc) >= 8 else acc
            acc_prd= acc[9:] if len(acc) > 9 else "01"

            # 시장가 / 지정가 구분
            ord_dvsn = "01" if price == 0 else "00"   # 01=시장가, 00=지정가

            try:
                resp = requests.post(
                    f"{self.base_url}/uapi/domestic-stock/v1/trading/order-cash",
                    headers=self._headers(tr_id),
                    json={
                        "CANO":         acc_no,
                        "ACNT_PRDT_CD": acc_prd,
                        "PDNO":         stock_code,
                        "ORD_DVSN":     ord_dvsn,
                        "ORD_QTY":      str(qty),
                        "ORD_UNPR":     str(price),
                    },
                    timeout=10
                )
                data = resp.json()

                if data.get("rt_cd") == "0":
                    self.daily_trades += 1
                    order_no  = data.get("output", {}).get("ODNO", "")
                    cur_price = price or self.get_current_price(stock_code)

                    # 포지션 추적
                    self._update_position(stock_code, order_type, qty, cur_price)

                    # 거래 로그 저장
                    self._save_log(stock_code, order_type, qty, cur_price, order_no, reason)

                    mode_str = "🔴실전" if self.is_real else "🟡모의"
                    emoji    = "📈" if order_type == "BUY" else "📉"
                    print(f"  {emoji} [{mode_str}] {order_type} {stock_code} "
                          f"{qty:,}주 @{cur_price:,}원 → 주문번호: {order_no}")
                    return {"success": True, "order_no": order_no, "price": cur_price}
                else:
                    print(f"  ⚠ KIS 주문 실패: {data.get('msg1', '')}")
                    return {"success": False, "msg": data.get("msg1", "")}

            except Exception as e:
                print(f"  ⚠ KIS 주문 오류: {e}")
                return {"success": False, "msg": str(e)}

    def _update_position(self, code, order_type, qty, price):
        if order_type == "BUY":
            if code in self.positions:
                old   = self.positions[code]
                total = old["qty"] + qty
                avg   = (old["avg_price"] * old["qty"] + price * qty) / total
                self.positions[code] = {
                    "qty": total, "avg_price": round(avg),
                    "buy_date": old["buy_date"]
                }
            else:
                self.positions[code] = {
                    "qty": qty, "avg_price": price,
                    "buy_date": datetime.today().strftime("%Y-%m-%d")
                }
        elif order_type == "SELL" and code in self.positions:
            new_qty = self.positions[code]["qty"] - qty
            if new_qty <= 0:
                del self.positions[code]
            else:
                self.positions[code]["qty"] = new_qty

    def _save_log(self, code, order_type, qty, price, order_no, reason):
        """거래 로그를 trades_KR_YYYYMMDD.json 에 누적 저장"""
        fname = datetime.today().strftime("trades_KR_%Y%m%d.json")
        fpath = os.path.join(TRADE_DIR, fname)
        logs  = []
        if os.path.exists(fpath):
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    logs = json.load(f)
            except Exception:
                pass
        logs.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type":      order_type,
            "code":      code,
            "qty":       qty,
            "price":     price,
            "amount":    qty * price,
            "order_no":  order_no,
            "reason":    reason,
            "mode":      "REAL" if self.is_real else "MOCK",
        })
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)

    # ── 시그널 실행 ──
    def execute_signals(self, df_signals: pd.DataFrame, total_capital: int) -> list:
        """
        스크리닝 매매 시그널 자동 실행
        ① 기존 포지션 손절·익절 점검
        ② 신규 매수 시그널 (강력매수·매수) 최대 5종목
        """
        if not self._is_configured():
            print("  ⚠ KIS 앱키 미설정 → 자동매매 중단")
            return []

        max_pos = self.cfg.get("max_position_pct", 5.0) / 100
        stop    = self.cfg.get("stop_loss_pct",    8.0) / 100
        tp      = self.cfg.get("take_profit_pct", 20.0) / 100
        executed = []

        # ① 포지션 손절·익절 점검
        print(f"\n  [KIS] 포지션 점검 ({len(self.positions)}개)...")
        for code, pos in list(self.positions.items()):
            cur = self.get_current_price(code)
            if cur <= 0:
                continue
            ret = (cur - pos["avg_price"]) / pos["avg_price"]

            if ret <= -stop:
                print(f"  🔴 손절: {code}  수익률 {ret*100:.1f}% → 전량매도")
                r = self.place_order(code, "SELL", pos["qty"],
                                     reason=f"손절 {ret*100:.1f}%")
                if r.get("success"):
                    executed.append({"action": "STOP_LOSS", "code": code,
                                     "ret_pct": round(ret * 100, 1)})

            elif ret >= tp:
                sell_qty = max(1, pos["qty"] // 2)
                print(f"  🟢 익절: {code}  수익률 {ret*100:.1f}% → {sell_qty}주 매도")
                r = self.place_order(code, "SELL", sell_qty,
                                     reason=f"익절 {ret*100:.1f}%")
                if r.get("success"):
                    executed.append({"action": "TAKE_PROFIT", "code": code,
                                     "ret_pct": round(ret * 100, 1)})

        # ② 신규 매수
        if "매매시그널" not in df_signals.columns:
            return executed

        buy_df = df_signals[
            df_signals["매매시그널"].str.contains("매수", na=False)
        ].head(5)
        print(f"  [KIS] 매수 시그널 {len(buy_df)}개 처리 중...")

        for code, row in buy_df.iterrows():
            str_code = str(code)
            if str_code in self.positions:
                continue   # 이미 보유 중

            cur = self.get_current_price(str_code)
            if cur <= 0:
                continue

            qty = max(1, int(total_capital * max_pos / cur))
            sig = str(row.get("매매시그널", ""))
            rsn = (f"{sig} | "
                   f"복합:{row.get('복합점수', 0):.1f} | "
                   f"100점:{row.get('100점_합계', 0):.0f} | "
                   f"괴리율:{row.get('괴리율(%)', 'N/A')}")

            r = self.place_order(str_code, "BUY", qty, reason=rsn)
            if r.get("success"):
                executed.append({"action": "BUY", "code": str_code,
                                  "qty": qty, "price": r.get("price")})
            time.sleep(0.5)   # API 과부하 방지

        print(f"  [KIS] 자동매매 완료: {len(executed)}건")
        return executed

    def reset_daily(self):
        """일일 거래 카운터 리셋 (매일 장 시작 전 호출)"""
        self.daily_trades = 0
        print(f"  [KIS] 일일 거래카운터 리셋 ({datetime.now().strftime('%Y-%m-%d')})")

    def status_summary(self) -> str:
        mode = "🔴실전" if self.is_real else "🟡모의"
        cfg_ok = "✅설정됨" if self._is_configured() else "❌미설정(kis_config.json 확인)"
        return (f"KIS API {mode} {cfg_ok} | "
                f"계좌:{self.cfg.get('account_no', '미설정')} | "
                f"오늘 거래:{self.daily_trades}회")


# ══════════════════════════════════════════════════════════
# NEW ④ 실시간 모니터링 & 텔레그램 알림
# ══════════════════════════════════════════════════════════
class MonitorEngine:
    """
    텔레그램 알림 + 성과 추적 엔진

    설정파일: telegram_config.json
    {
        "bot_token": "YOUR_BOT_TOKEN",
        "chat_id":   "YOUR_CHAT_ID"
    }

    봇 설정 방법:
      1. 텔레그램에서 @BotFather → /newbot → 토큰 복사
      2. 봇에게 메시지를 먼저 보낸 뒤
         https://api.telegram.org/bot<TOKEN>/getUpdates 에서 chat_id 확인
    """

    def __init__(self, config_path: str = None):
        self.config_path  = config_path or os.path.join(BASE_DIR, "telegram_config.json")
        self.bot_token    = ""
        self.chat_id      = ""
        self.enabled      = False
        self.perf_history = []
        self._load_config()
        self._load_perf_history()

    def _load_config(self):
        if not os.path.exists(self.config_path):
            self._create_default_config()
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            self.bot_token = cfg.get("bot_token", "")
            self.chat_id   = cfg.get("chat_id",   "")
            self.enabled   = bool(self.bot_token and
                                  self.bot_token != "YOUR_BOT_TOKEN")
            print(f"  [텔레그램] {'✅ 활성' if self.enabled else '❌ 비활성 (설정 필요)'}")
        except Exception as e:
            print(f"  ⚠ 텔레그램 설정 오류: {e}")

    def _create_default_config(self):
        default = {"bot_token": "YOUR_BOT_TOKEN", "chat_id": "YOUR_CHAT_ID"}
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        print(f"  [텔레그램] 설정파일 생성: {self.config_path}")

    def _load_perf_history(self):
        path = os.path.join(LOG_DIR, "perf_history.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    self.perf_history = json.load(f)
            except Exception:
                self.perf_history = []

    def _save_perf_history(self):
        path = os.path.join(LOG_DIR, "perf_history.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.perf_history[-365:], f, ensure_ascii=False, indent=2)

    def send(self, text: str) -> bool:
        """텔레그램 메시지 전송"""
        if not self.enabled:
            print(f"  [텔레그램 비활성] {text[:80]}...")
            return False
        url = f"https://api.telegram.org/bot{self.bot_token}/sendMessage"
        try:
            r = requests.post(url, json={
                "chat_id": self.chat_id, "text": text, "parse_mode": "HTML"
            }, timeout=10)
            return r.status_code == 200
        except Exception as e:
            print(f"  ⚠ 텔레그램 전송 실패: {e}")
            return False

    def notify_screening(self, df_top: pd.DataFrame) -> None:
        """스크리닝 결과 알림 (상위 10종목)"""
        today = datetime.today().strftime("%Y-%m-%d")
        lines = [
            f"📊 <b>퀀트 스크리너 {VERSION} [KR]</b>",
            f"🗓 {today}",
            "━━━━━━━━━━━━━━━━━━━━",
        ]
        for rank, (code, row) in enumerate(df_top.head(10).iterrows(), 1):
            name   = str(row.get("종목명", ""))[:8]
            score  = _safe(row.get("강화복합점수") or row.get("복합점수"), 0)
            pt100  = _safe(row.get("100점_합계"), 0)
            sig    = str(row.get("매매시그널", "─"))
            upside = _safe(row.get("괴리율(%)"))

            if "강력매수" in sig:   em = "🔵"
            elif "■■ 매수" in sig:  em = "🟢"
            elif "관심" in sig:     em = "🟡"
            elif "매도" in sig:     em = "🔴"
            else:                   em = "⬜"

            up_str = f"({upside:+.0f}%)" if upside is not None else ""
            lines.append(
                f"{rank:2d}. {em} <b>{name}</b>({code}) "
                f"복합:{score:.0f} 100점:{pt100:.0f} {up_str}"
            )

        lines.append("━━━━━━━━━━━━━━━━━━━━")

        # 강력매수 강조
        if "매매시그널" in df_top.columns:
            sb = df_top[df_top["매매시그널"].str.contains("강력매수", na=False)]
            if not sb.empty:
                lines.append(f"🔵 <b>강력매수 {len(sb)}종목</b>")
                for code, row in sb.iterrows():
                    name = str(row.get("종목명", ""))[:8]
                    tgt  = _safe(row.get("1차목표가"))
                    stp  = _safe(row.get("손절선"))
                    if tgt and stp:
                        lines.append(f"   {name}({code}) → "
                                     f"목표:{int(tgt):,}원 / 손절:{int(stp):,}원")
                    else:
                        lines.append(f"   {name}({code})")

        self.send("\n".join(lines))

    def notify_positions(self, trader: "KISAutoTrader") -> None:
        """포지션 현황 알림"""
        if not trader.positions:
            self.send("📋 현재 보유 포지션 없음")
            return
        lines = [f"📋 <b>포지션 현황</b> ({datetime.now().strftime('%H:%M')})",
                 "━━━━━━━━━━━━━━━━━━━━"]
        total_pnl = 0
        for code, pos in trader.positions.items():
            cur = trader.get_current_price(code)
            if cur > 0:
                ret     = (cur - pos["avg_price"]) / pos["avg_price"] * 100
                pnl_won = (cur - pos["avg_price"]) * pos["qty"]
                total_pnl += pnl_won
                em = "📈" if ret >= 0 else "📉"
                lines.append(f"{em} {code}: {pos['qty']}주 "
                             f"@{pos['avg_price']:,}→{cur:,}원 "
                             f"({ret:+.1f}%, {pnl_won:+,.0f}원)")
        lines.append("━━━━━━━━━━━━━━━━━━━━")
        lines.append(f"💰 총 평가손익: {total_pnl:+,.0f}원")
        self.send("\n".join(lines))

    def notify_trade(self, executed: list) -> None:
        """자동매매 실행 결과 알림"""
        if not executed:
            return
        lines = ["🤖 <b>자동매매 실행 결과</b>"]
        for e in executed:
            action = e.get("action", "")
            code   = e.get("code", "")
            if action == "BUY":
                lines.append(f"  📈 매수: {code} "
                             f"{e.get('qty',0)}주 @{e.get('price',0):,}원")
            elif action == "STOP_LOSS":
                lines.append(f"  🔴 손절: {code} ({e.get('ret_pct',0):+.1f}%)")
            elif action == "TAKE_PROFIT":
                lines.append(f"  🟢 익절: {code} ({e.get('ret_pct',0):+.1f}%)")
        self.send("\n".join(lines))

    def notify_signal_change(self, df_today: pd.DataFrame,
                              df_prev: Optional[pd.DataFrame]) -> None:
        """전일 대비 시그널 변경 감지"""
        if df_prev is None or df_prev.empty:
            return
        changes = []
        for code in df_today.index:
            if code not in df_prev.index:
                continue
            s_now  = str(df_today.loc[code, "매매시그널"]
                         if "매매시그널" in df_today.columns else "")
            s_prev = str(df_prev.loc[code, "매매시그널"]
                         if "매매시그널" in df_prev.columns else "")
            if s_now and s_now != s_prev:
                name = str(df_today.loc[code, "종목명"]
                           if "종목명" in df_today.columns else code)
                changes.append(f"  {name}({code}): {s_prev} → <b>{s_now}</b>")
        if changes:
            self.send("⚡ <b>시그널 변경</b>\n" + "\n".join(changes[:10]))

    def track(self, df_top: pd.DataFrame) -> None:
        """성과 히스토리 기록 + 주간 리포트 (월요일)"""
        top10 = df_top.index[:10].tolist()
        scores = [_safe(df_top.loc[c, "강화복합점수"]
                        if "강화복합점수" in df_top.columns
                        else df_top.loc[c, "복합점수"], 0)
                  for c in top10 if c in df_top.index]
        entry = {
            "date":     datetime.today().strftime("%Y-%m-%d"),
            "top10":    [str(c) for c in top10],
            "avg_score":round(float(np.mean(scores)), 1) if scores else 0,
            "강력매수": int(df_top["매매시그널"].str.contains("강력매수", na=False).sum())
                        if "매매시그널" in df_top.columns else 0,
        }
        self.perf_history.append(entry)
        self._save_perf_history()

        if datetime.today().weekday() == 0:   # 월요일 → 주간 리포트
            self._weekly_report()

    def _weekly_report(self):
        week_ago = (datetime.today() - timedelta(days=7)).strftime("%Y-%m-%d")
        recent   = [e for e in self.perf_history if e["date"] >= week_ago]
        if not recent:
            return
        avg_score  = np.mean([e.get("avg_score", 0) for e in recent])
        avg_strong = np.mean([e.get("강력매수", 0) for e in recent])
        self.send(
            f"📅 <b>주간 리포트 [KR]</b>\n"
            f"기간: {week_ago} ~ {datetime.today().strftime('%Y-%m-%d')}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"평균 복합점수: {avg_score:.1f}\n"
            f"일평균 강력매수: {avg_strong:.1f}개\n"
            f"스크리닝 실행: {len(recent)}회"
        )

    def load_yesterday(self) -> Optional[pd.DataFrame]:
        """전일 매매신호 JSON 로드"""
        yesterday = (datetime.today() - timedelta(days=1)).strftime("%Y%m%d")
        fpath = os.path.join(BASE_DIR, f"매매신호_KR_{yesterday}.json")
        if not os.path.exists(fpath):
            return None
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            rows = {s["code"]: {
                "종목명":     s.get("name", ""),
                "매매시그널": s.get("trading", {}).get("signal", ""),
                "복합점수":   s.get("composite_score", 0),
            } for s in data.get("top_stocks", [])}
            return pd.DataFrame.from_dict(rows, orient="index") if rows else None
        except Exception:
            return None


# ══════════════════════════════════════════════════════════
# v35 엔진 동적 임포트 (한국 전용 함수만 로드)
# ══════════════════════════════════════════════════════════
_v35_path = os.path.join(BASE_DIR, "quant_screener_v35.py")
HAS_V35   = False

if os.path.exists(_v35_path):
    import importlib.util
    _spec = importlib.util.spec_from_file_location("quant_v35", _v35_path)
    _v35  = importlib.util.module_from_spec(_spec)
    try:
        _spec.loader.exec_module(_v35)
        DartClient             = _v35.DartClient
        fetch_kr_universe      = _v35.fetch_kr_universe
        fetch_kr_all_data      = _v35.fetch_kr_all_data
        compute_scores         = _v35.compute_scores
        apply_hard_filter      = _v35.apply_hard_filter
        build_excel            = _v35.build_excel
        save_signal_json       = _v35.save_signal_json
        print_summary          = _v35.print_summary
        load_dart_key          = _v35.load_dart_key
        HAS_V35 = True
        print(f"  ✅ v35 엔진 로드 완료 (한국 전용)")
    except Exception as e:
        print(f"  ⚠ v35 로드 실패: {e}")
else:
    print(f"  ⚠ quant_screener_v35.py 없음")


# ══════════════════════════════════════════════════════════
# 파일명 래퍼 — 항상 날짜+시분초로 저장 (중복 방지)
#
#  v35 build_excel / save_signal_json 는 내부적으로
#  %Y%m%d (날짜만) 파일명을 사용해 같은 날 실행하면 덮어씀.
#  v36 에서는 %Y%m%d_%H%M%S 형식을 강제해 항상 새 파일 생성.
#
#  파일명 예시:
#    quant_KR_20260618_091530.xlsx   ← 9시15분30초 실행분
#    quant_KR_20260618_143022.xlsx   ← 14시30분22초 실행분
#    매매신호_KR_20260618_091530.json
#    backtest_KR_20260618_091530.xlsx
# ══════════════════════════════════════════════════════════

def _ts() -> str:
    """현재 실행 타임스탬프 (날짜+시분초), 한 번 생성 후 모듈 변수로 고정"""
    return datetime.now().strftime("%Y%m%d_%H%M%S")

# 실행 시작 시각을 모듈 수준에서 한 번만 기록
# → 같은 실행에서 엑셀과 JSON 파일명이 동일한 타임스탬프를 가짐
_RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")


def build_excel_v36(df_all, df_top, label, args, output_dir=None):
    """
    v35 build_excel을 호출한 뒤 파일명을 날짜+시분초로 변경.
    - v35 내부에서는 quant_{label}_%Y%m%d.xlsx 로 저장됨
    - 저장 직후 quant_{label}_%Y%m%d_%H%M%S.xlsx 로 rename
    """
    if not HAS_V35:
        return None
    out_dir = output_dir or BASE_DIR

    # v35 원본 저장 (날짜만 파일명)
    result_path = build_excel(df_all, df_top, label, args, out_dir)
    if not result_path or not os.path.exists(result_path):
        return result_path

    # 날짜만 파일명 → 날짜+시분초 파일명으로 rename
    date_only = datetime.now().strftime(f"quant_{label}_%Y%m%d.xlsx")
    new_name  = f"quant_{label}_{_RUN_TS}.xlsx"
    new_path  = os.path.join(out_dir, new_name)

    # 혹시 target 파일이 이미 있으면 삭제 후 rename
    if os.path.exists(new_path):
        os.remove(new_path)
    try:
        os.rename(result_path, new_path)
        print(f"  📁 파일명 확정: {new_name}")
    except Exception as e:
        print(f"  ⚠ 파일명 변경 실패({e}) → 원본 유지: {result_path}")
        return result_path
    return new_path


def save_signal_json_v36(df_top, label, args, output_dir=None):
    """
    v35 save_signal_json을 호출한 뒤 파일명을 날짜+시분초로 변경.
    - v35 내부: 매매신호_{label}_%Y%m%d.json
    - rename 후: 매매신호_{label}_%Y%m%d_%H%M%S.json
    """
    if not HAS_V35:
        return
    out_dir = output_dir or BASE_DIR

    save_signal_json(df_top, label, args, out_dir)

    old_name = datetime.now().strftime(f"매매신호_{label}_%Y%m%d.json")
    old_path = os.path.join(out_dir, old_name)
    new_name = f"매매신호_{label}_{_RUN_TS}.json"
    new_path = os.path.join(out_dir, new_name)

    if not os.path.exists(old_path):
        return
    if os.path.exists(new_path):
        os.remove(new_path)
    try:
        os.rename(old_path, new_path)
        print(f"  📁 파일명 확정: {new_name}")
    except Exception as e:
        print(f"  ⚠ 파일명 변경 실패({e}) → 원본 유지: {old_path}")


# ══════════════════════════════════════════════════════════
# 인수 파서 (한국 전용으로 단순화)
# ══════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser(
        description=f"퀀트 주식 스크리너 {VERSION} — 한국 KOSPI/KOSDAQ 전용"
    )
    # 스크리닝 파라미터
    p.add_argument("--top",          type=int,   default=20,    help="상위 종목 수")
    p.add_argument("--min-roe",      type=float, default=8.0)
    p.add_argument("--max-pbr",      type=float, default=4.0)
    p.add_argument("--max-debt",     type=float, default=150.0)
    p.add_argument("--min-ic",       type=float, default=2.0)
    p.add_argument("--min-per",      type=float, default=0.0)
    p.add_argument("--max-per",      type=float, default=999.0)
    # 팩터 가중치
    p.add_argument("--w-value",      type=float, default=20.0)
    p.add_argument("--w-mom",        type=float, default=20.0)
    p.add_argument("--w-div",        type=float, default=5.0)
    p.add_argument("--w-quality",    type=float, default=15.0)
    p.add_argument("--w-growth",     type=float, default=15.0)
    p.add_argument("--w-prof",       type=float, default=15.0)
    p.add_argument("--w-stability",  type=float, default=10.0)
    # 실행 옵션
    p.add_argument("--output-dir",   type=str,   default=BASE_DIR)
    p.add_argument("--no-cache",     action="store_true", help="캐시 무시")
    p.add_argument("--auto",         action="store_true", help="비대화형 자동 실행")
    p.add_argument("--scope",        type=str,   default="2",
                   choices=["1","2","3"], help="1=100개 2=300개 3=전체")
    # v36 신규
    p.add_argument("--backtest",     action="store_true", help="백테스트 전용")
    p.add_argument("--trade",        action="store_true", help="KIS 자동매매 실행 (한국투자증권 REST API)")
    p.add_argument("--monitor",      action="store_true", help="모니터링 전용")
    p.add_argument("--factor-model", action="store_true", help="Z-score 팩터모델 적용")
    p.add_argument("--no-telegram",  action="store_true", help="텔레그램 알림 비활성")
    p.add_argument("--bt-start",     type=str,   default="2020-01-01")
    p.add_argument("--bt-end",       type=str,   default="")
    p.add_argument("--bt-is",        type=int,   default=12, help="IS 개월수")
    p.add_argument("--bt-oos",       type=int,   default=3,  help="OOS 개월수")
    return p.parse_args()


# ══════════════════════════════════════════════════════════
# 실행 함수
# ══════════════════════════════════════════════════════════
def _run_backtest(args):
    print("\n" + "=" * 65)
    print(f"  📊 백테스트 모드 [{args.bt_start} ~ {args.bt_end or '오늘'}]")
    print("=" * 65)
    if not HAS_V35:
        print("  ⚠ quant_screener_v35.py 필요")
        return

    tickers, _, _ = fetch_kr_universe()
    bt = BacktestEngine(
        universe_codes=tickers[:200],
        start=args.bt_start,
        end=args.bt_end or None,
        is_months=args.bt_is,
        oos_months=args.bt_oos,
        top_n=args.top,
    )
    price_df = bt.load_price_data()
    if price_df.empty:
        print("  ⚠ 가격 데이터 없음")
        return

    result = bt.run_walkforward(price_df)
    if result:
        fpath = bt.save_to_excel(result)
        if not args.no_telegram:
            m = MonitorEngine()
            s = result.get("summary", {})
            m.send(
                f"📊 <b>백테스트 결과 [KR]</b>\n"
                f"CAGR: {s.get('CAGR(%)',0):+.1f}% | "
                f"Sharpe: {s.get('Sharpe_OOS',0):.2f}\n"
                f"MDD: {s.get('MDD(%)',0):.1f}% | "
                f"승률: {s.get('승률(%)',0):.1f}%\n"
                f"{s.get('과적합판단','')}"
            )


def _run_monitor(args):
    print("\n  [모니터] 모니터링 전용 실행")
    monitor = MonitorEngine()
    trader  = KISAutoTrader()
    monitor.notify_positions(trader)
    monitor.send(f"ℹ {trader.status_summary()}")


def _run_scan(args, scope: str):
    """한국 KOSPI/KOSDAQ 스크리닝 + 신규 엔진 통합"""
    if not HAS_V35:
        print("  ⚠ quant_screener_v35.py 없음")
        return

    kr_size = {"1": 100, "2": 300, "3": None}[scope]

    monitor = None if args.no_telegram else MonitorEngine()
    trader  = KISAutoTrader() if args.trade else None
    factor  = FactorModel(sector_neutral=True) if args.factor_model else None

    print("\n" + "─" * 55)
    print("  [한국 KOSPI/KOSDAQ 스크리닝]")
    print("─" * 55)

    # v35 데이터 수집
    dart_key = load_dart_key()
    dart     = DartClient(dart_key)
    tickers, names, markets = fetch_kr_universe()
    df_raw = fetch_kr_all_data(
        dart, tickers, names, markets, kr_size,
        no_cache=getattr(args, "no_cache", False)
    )
    df_raw = compute_scores(df_raw, args)

    # Z-score 팩터 모델 적용
    if factor:
        print("  [팩터모델] Z-score 섹터중립화 적용 중...")
        df_raw = factor.enhance_scores(df_raw)

    df_filtered = apply_hard_filter(df_raw, args)
    if df_filtered.empty:
        print("  ⚠ 필터 통과 종목 없음")
        return

    # 강화복합점수 있으면 재정렬
    sort_col = "강화복합점수" if "강화복합점수" in df_filtered.columns else "복합점수"
    df_filtered = df_filtered.sort_values(sort_col, ascending=False)
    df_top      = df_filtered.head(args.top)

    print(f"  📊 통과: {len(df_filtered)}개 / 전체 {len(df_raw)}개 → 상위 {args.top}개")
    print_summary(df_top, "한국 KOSPI/KOSDAQ", args)

    # 엑셀·JSON 저장 (날짜+시분초 파일명 — 실행마다 새 파일 생성)
    build_excel_v36(df_raw, df_top, "KR", args, args.output_dir)
    save_signal_json_v36(df_top, "KR", args, args.output_dir)

    # 텔레그램
    if monitor:
        df_prev = monitor.load_yesterday()
        monitor.notify_signal_change(df_top, df_prev)
        monitor.notify_screening(df_top)
        monitor.track(df_top)

    # KIS 자동매매
    if trader:
        print("\n  [KIS] 자동매매 실행...")
        bal       = trader.get_balance()
        total_cap = bal.get("순자산", 10_000_000)
        executed  = trader.execute_signals(df_top, total_cap)
        if monitor and executed:
            monitor.notify_trade(executed)

    print("\n  ✅ 완료")
    print("  📌 신규 기능:")
    print(f"     팩터모델:    {'✅' if factor else '─ (--factor-model)'}")
    print(f"     KIS 자동매매: {'✅' if trader else '─ (--trade / kis_config.json 설정)'}")
    print(f"     텔레그램:    {'✅' if (monitor and monitor.enabled) else '─ (telegram_config.json 설정)'}")


# ══════════════════════════════════════════════════════════
# 대화형 메뉴
# ══════════════════════════════════════════════════════════
def interactive_menu():
    print()
    print("  ┌──────────────────────────────────────────────────────────┐")
    print(f"  │  퀀트 주식 스크리너 {VERSION}  (한국 KOSPI/KOSDAQ 전용)  │")
    print("  │  v35 전체 + 백테스트 + Z팩터 + KIS자동매매 + 텔레그램  │")
    print("  └──────────────────────────────────────────────────────────┘")
    print()
    cache_show_status()
    cache_clear(days_old=8)
    print()

    print("  [실행 모드]")
    print("  1. 스크리닝  (v35 전체 + 팩터모델 강화)")
    print("  2. 스크리닝 + KIS 자동매매  (한국투자증권 REST API, 64bit)")
    print("  3. 백테스트  (Walk-forward IS/OOS)")
    print("  4. 모니터링  (텔레그램 포지션 현황)")
    print("  5. 설정 안내  (KIS API / 텔레그램)")
    while True:
        mode = input("\n  선택 (1~5): ").strip()
        if mode in ("1","2","3","4","5"):
            break
        print("  1~5 중 하나를 입력하세요.")

    if mode == "5":
        print()
        print("  ── KIS API 설정 (한국투자증권, 64bit Python 지원) ──")
        print(f"  → 설정파일: {os.path.join(BASE_DIR, 'kis_config.json')}")
        print("  → https://apiportal.koreainvestment.com 에서 앱키 무료 발급")
        print("  → app_key / app_secret / account_no 입력 후 저장")
        print("  → is_real: false (모의) → true (실전) 로 변경하면 실전 주문")
        print()
        print("  ── 텔레그램 설정 ──")
        print(f"  → 설정파일: {os.path.join(BASE_DIR, 'telegram_config.json')}")
        print("  → @BotFather → /newbot → 토큰 발급")
        print("  → https://api.telegram.org/bot<TOKEN>/getUpdates → chat_id 확인")
        sys.exit(0)

    if mode == "3":
        return "backtest", None
    if mode == "4":
        return "monitor", None

    print()
    print("  [스크리닝 범위]")
    print("  1. 빠른 테스트  (KOSPI+KOSDAQ 100개,  약 10~15분)")
    print("  2. 중간 범위    (KOSPI+KOSDAQ 300개,  약 30~40분)")
    print("  3. 전체 스크리닝 (전 종목,              수 시간)")
    while True:
        scope = input("\n  선택 (1/2/3): ").strip()
        if scope in ("1","2","3"):
            break

    factor_yn = input("\n  Z-score 팩터모델 적용? (y/n, 기본 n): ").strip().lower()
    trade_yn  = ("y" if mode == "2"
                 else input("  KIS 자동매매 실행? (y/n, 기본 n): ").strip().lower())

    return scope, {"factor_model": factor_yn == "y", "trade": trade_yn == "y"}


# ══════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════
def main():
    print()
    print("=" * 65)
    print(f"  퀀트 스크리너 {VERSION}  —  한국 KOSPI/KOSDAQ 전용")
    print("  v35 전체 + 백테스트 + Z팩터 + KIS자동매매(64bit) + 텔레그램")
    print("=" * 65)

    args = parse_args()

    # 단독 모드
    if args.backtest:
        _run_backtest(args)
        return
    if args.monitor:
        _run_monitor(args)
        return

    if args.auto:
        # 비대화형 자동 실행
        scope = args.scope
        print(f"\n  🤖 자동 실행  [{datetime.now().strftime('%Y-%m-%d %H:%M')}]")
        cache_show_status()
        cache_clear(days_old=8)
    else:
        # 대화형 메뉴
        result = interactive_menu()
        if result[0] == "backtest":
            _run_backtest(args)
            return
        if result[0] == "monitor":
            _run_monitor(args)
            return

        scope, opts = result
        opts = opts or {}
        if opts.get("factor_model"):
            args.factor_model = True
        if opts.get("trade"):
            args.trade = True

    _run_scan(args, scope)


if __name__ == "__main__":
    main()
